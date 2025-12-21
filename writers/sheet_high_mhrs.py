"""
High Man-Hours Sheet Module
FIXED: Removed coefficient columns, only shows Base Hours
Also adds red highlighting for blank SEQ rows
"""

import pandas as pd
from openpyxl.styles import PatternFill
from utils.time_utils import hours_to_hhmm
from core.config import SEQ_NO_COLUMN, TITLE_COLUMN


def create_high_mhrs_sheet(writer, report_data):
    """
    Create the High Man-Hours Tasks sheet.
    FIXED: Only shows Base Hours (no coefficient or adjusted hours)
    """
    high_mhrs_df = report_data['high_mhrs_tasks'].copy()

    if len(high_mhrs_df) == 0:
        df = pd.DataFrame([['No tasks found with planned man-hours exceeding the threshold']])
        df.to_excel(writer, sheet_name='High Man-Hours Tasks', index=False, header=False)
        return

    # Add HH:MM formatted column (ONLY Base Hours)
    high_mhrs_df['Base Mhrs'] = high_mhrs_df['Base Hours'].apply(hours_to_hhmm)

    # Select and order columns (NO coefficient or adjusted hours)
    columns_to_export = build_export_columns(high_mhrs_df)
    export_df = high_mhrs_df[columns_to_export]

    # Write to Excel
    export_df.to_excel(writer, sheet_name='High Man-Hours Tasks', index=False)

    # Get the worksheet
    worksheet = writer.sheets['High Man-Hours Tasks']
    worksheet.auto_filter.ref = worksheet.dimensions

    # Auto-adjust column widths
    adjust_column_widths(writer, 'High Man-Hours Tasks', export_df)

    # Add red highlighting for blank SEQ rows
    highlight_blank_seq_rows(worksheet, export_df)


def build_export_columns(df):
    """Build the list of columns to export (NO coefficients)."""
    columns_to_export = []

    if SEQ_NO_COLUMN in df.columns:
        columns_to_export.append(SEQ_NO_COLUMN)

    if TITLE_COLUMN in df.columns:
        columns_to_export.append(TITLE_COLUMN)

    if 'Task ID' in df.columns:
        columns_to_export.append('Task ID')

    # ONLY Base Hours (no coefficient, no adjusted hours)
    columns_to_export.append('Base Mhrs')

    return columns_to_export


def highlight_blank_seq_rows(worksheet, df):
    """
    Add red highlighting to rows with blank SEQ values.

    Args:
        worksheet: openpyxl worksheet object
        df: DataFrame that was written to the sheet
    """
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # Find SEQ column index
    seq_col_idx = None
    for idx, col in enumerate(df.columns):
        if col == SEQ_NO_COLUMN:
            seq_col_idx = idx
            break

    if seq_col_idx is None:
        return  # SEQ column not found

    # Check each data row (starting from row 2, after header)
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        seq_value = row[SEQ_NO_COLUMN]

        # Check if SEQ is blank/empty
        if pd.isna(seq_value) or str(seq_value).strip() == '':
            # Highlight entire row
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = red_fill


def adjust_column_widths(writer, sheet_name, df, max_width=50):
    """Auto-adjust column widths."""
    worksheet = writer.sheets[sheet_name]

    for idx, col in enumerate(df.columns):
        max_length = max(
            df[col].astype(str).apply(len).max(),
            len(str(col))
        )
        worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, max_width)