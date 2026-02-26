"""
New Task IDs Sheet Module
UPDATED: Added 'Description' column sourced from forecast.event_display (TITLE_COLUMN)
FIXED: Red highlighting for blank SEQ rows
"""

import pandas as pd
from openpyxl.styles import PatternFill
from core.config import TITLE_COLUMN


def create_new_task_ids_sheet(writer, report_data):
    """
    Create the New Task IDs sheet with SEQ numbers, Task IDs, and Descriptions.

    Columns: SEQ | New Task ID | Description
    """
    new_task_ids_df = report_data['new_task_ids_with_seq']

    if len(new_task_ids_df) == 0:
        df = pd.DataFrame([['No new task IDs found - all task IDs match reference']])
        df.to_excel(writer, sheet_name='New Task IDs', index=False, header=False)
        return

    # Filter out None / 'nan' Task IDs
    filtered_df = filter_valid_task_ids(new_task_ids_df)

    if len(filtered_df) == 0:
        df = pd.DataFrame([['No new task IDs found - all task IDs match reference']])
        df.to_excel(writer, sheet_name='New Task IDs', index=False, header=False)
        return

    # Build the export DataFrame with consistent columns
    export_df = _build_export_df(filtered_df)

    # Write to Excel
    export_df.to_excel(writer, sheet_name='New Task IDs', index=False)

    # Worksheet formatting
    worksheet = writer.sheets['New Task IDs']
    worksheet.auto_filter.ref = worksheet.dimensions

    adjust_column_widths(writer, 'New Task IDs', export_df)
    highlight_blank_seq_rows(worksheet, export_df)


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

def _build_export_df(df):
    """
    Build the final export DataFrame.

    Input df may have columns: [SEQ_NO_COLUMN, 'Task ID', TITLE_COLUMN]
    Output always has columns:  ['SEQ', 'New Task ID', 'Description']

    If TITLE_COLUMN is missing from df (e.g., older data), Description is blank.
    """
    # Map source columns → display columns
    col_map = {}

    # Identify the SEQ column (first column in incoming df, or by name)
    src_cols = list(df.columns)

    # Find SEQ column
    from core.config import SEQ_NO_COLUMN
    seq_col = SEQ_NO_COLUMN if SEQ_NO_COLUMN in src_cols else src_cols[0]

    col_map[seq_col] = 'SEQ'
    col_map['Task ID'] = 'New Task ID'

    # Description from TITLE_COLUMN if present
    has_description = TITLE_COLUMN in src_cols

    out = pd.DataFrame()
    out['SEQ'] = df[seq_col].fillna('').astype(str)
    out['New Task ID'] = df['Task ID'].fillna('').astype(str)
    out['Description'] = df[TITLE_COLUMN].fillna('').astype(str) if has_description else ''

    return out


def filter_valid_task_ids(df):
    """Filter out rows where Task ID is None / 'nan'."""
    return df[
        df['Task ID'].notna() &
        (df['Task ID'].astype(str).str.lower() != 'nan') &
        (df['Task ID'].astype(str).str.strip() != '')
    ].copy()


def highlight_blank_seq_rows(worksheet, df):
    """Add red highlighting to rows where SEQ is blank/empty."""
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        seq_value = row.get('SEQ', '')
        if pd.isna(seq_value) or str(seq_value).strip() in ('', 'nan'):
            for col_idx in range(1, len(df.columns) + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = red_fill


def adjust_column_widths(writer, sheet_name, df):
    """Auto-adjust column widths with sensible maximums."""
    worksheet = writer.sheets[sheet_name]

    max_widths = {
        'SEQ': 20,
        'New Task ID': 35,
        'Description': 80,
    }

    for idx, col in enumerate(df.columns):
        max_data_len = df[col].fillna('').astype(str).apply(len).max()
        col_width = min(
            max(max_data_len, len(str(col))) + 2,
            max_widths.get(col, 40)
        )
        worksheet.column_dimensions[chr(65 + idx)].width = col_width


# ─────────────────────────────────────────────────────────────────────────────
# Public utility helpers (used by other modules)
# ─────────────────────────────────────────────────────────────────────────────

def count_new_task_ids(new_task_ids_df):
    """Return the count of valid new task IDs."""
    return len(filter_valid_task_ids(new_task_ids_df))


def get_new_task_ids_summary(new_task_ids_df):
    """Return a summary dict for the new task IDs."""
    filtered_df = filter_valid_task_ids(new_task_ids_df)
    if len(filtered_df) == 0:
        return {'total_new_ids': 0, 'unique_seqs': 0}
    return {
        'total_new_ids': len(filtered_df),
        'unique_seqs': filtered_df.iloc[:, 0].nunique(),
    }