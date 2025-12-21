"""
Tool Control Sheet Module
FIXED: Added red highlighting for blank SEQ rows
"""

import pandas as pd
from openpyxl.styles import PatternFill


def create_tool_control_sheet(writer, report_data):
    """
    Create the Tool Control sheet showing tools/spares with zero availability.
    FIXED: Added red highlighting for blank SEQ rows.
    """
    tool_issues_df = report_data.get('tool_control_issues', pd.DataFrame())

    if len(tool_issues_df) == 0:
        df = pd.DataFrame([['All tools and spares have adequate availability (quantity > 0)']])
        df.to_excel(writer, sheet_name='Tool Control', index=False, header=False)
        return

    # Write to Excel with headers
    tool_issues_df.to_excel(writer, sheet_name='Tool Control', index=False)

    # Get the worksheet
    worksheet = writer.sheets['Tool Control']
    worksheet.auto_filter.ref = worksheet.dimensions

    # Auto-adjust column widths
    adjust_column_widths(writer, 'Tool Control', tool_issues_df)

    # Add red highlighting for blank SEQ rows
    highlight_blank_seq_rows(worksheet, tool_issues_df)


def highlight_blank_seq_rows(worksheet, df):
    """
    Add red highlighting to rows with blank SEQ values.

    Args:
        worksheet: openpyxl worksheet object
        df: DataFrame that was written to the sheet
    """
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # Find SEQ column index
    seq_col_idx = None
    for idx, col in enumerate(df.columns):
        if col == 'SEQ':
            seq_col_idx = idx
            break

    if seq_col_idx is None:
        return  # SEQ column not found

    # Check each data row (starting from row 2, after header)
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        seq_value = row['SEQ']

        # Check if SEQ is blank/empty
        if pd.isna(seq_value) or str(seq_value).strip() == '':
            # Highlight entire row
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = red_fill


def adjust_column_widths(writer, sheet_name, df):
    """Auto-adjust column widths for better readability."""
    worksheet = writer.sheets[sheet_name]

    column_max_widths = {
        'Tool/Spare Name': 70,
        'Part Number': 25,
        'Task ID': 30,
        'SEQ': 20,
        'Type': 20
    }

    for idx, col in enumerate(df.columns):
        max_length = max(
            df[col].astype(str).apply(len).max(),
            len(str(col))
        )
        max_width = column_max_widths.get(col, 20)
        worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, max_width)


def get_tool_control_summary(tool_issues_df):
    """Generate a summary of tool control issues."""
    if len(tool_issues_df) == 0:
        return {
            'total_issues': 0,
            'tools': 0,
            'spares': 0,
            'unique_parts': 0,
            'affected_seqs': 0
        }

    return {
        'total_issues': len(tool_issues_df),
        'tools': len(tool_issues_df[tool_issues_df['Type'] == 'Tool']),
        'spares': len(tool_issues_df[tool_issues_df['Type'] == 'Spare']),
        'unique_parts': tool_issues_df['Part Number'].nunique() if 'Part Number' in tool_issues_df.columns else 0,
        'affected_seqs': tool_issues_df['SEQ'].nunique() if 'SEQ' in tool_issues_df.columns else 0
    }


def format_tool_control_message():
    """Format the message for when no tool control issues are found."""
    return 'All tools and spares have adequate availability (quantity > 0)'


def count_tool_issues_by_type(tool_issues_df):
    """Count tool issues by type (Tool vs Spare)."""
    if len(tool_issues_df) == 0:
        return {'Tool': 0, 'Spare': 0, 'Unknown': 0}

    type_counts = tool_issues_df['Type'].value_counts().to_dict()

    result = {
        'Tool': type_counts.get('Tool', 0),
        'Spare': type_counts.get('Spare', 0),
        'Unknown': type_counts.get('Unknown', 0)
    }

    return result